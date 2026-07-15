from flask import Flask, render_template, request, jsonify, Response, flash, session, url_for, redirect
import re
import cv2
import pandas as pd
import numpy as np
import os
import tempfile
import pytesseract
from pytesseract import Output
from PIL import Image
import fitz
from io import BytesIO
from werkzeug.datastructures import FileStorage
import base64
from collections import defaultdict
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.utils import secure_filename
from sqlalchemy.dialects.mysql import LONGBLOB, JSON
from sqlalchemy import LargeBinary, func, and_
import math
from datetime import datetime
import openpyxl
import imghdr
from sqlalchemy.orm.attributes import flag_modified
from sqlalchemy.orm import joinedload
from sqlalchemy import text
import time
import random
import zipfile
from xml.etree import ElementTree as ET
import io
from flask_mail import Mail, Message

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

NS = {
    "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships"
}
DISPIMG_RE = re.compile(r'DISPIMG\("([^"]+)"')

EXT_TO_MIME = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "gif": "image/gif",
    "bmp": "image/bmp",
}

app = Flask(__name__)
app.secret_key='7527488f33cd3a731909c7d6e8aa8194d85e893f02a7d8548cb4b16aec47f64f'

# Mail Configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'sonalikajewellers2021@gmail.com'
app.config['MAIL_PASSWORD'] = 'amag epeg aaez mklf'
app.config['MAIL_DEFAULT_SENDER'] = 'sonalikajewellers2021@gmail.com'

mail = Mail(app)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://sonalika:1234@localhost/sonalika'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config["MAX_CONTENT_LENGTH"] = 1024 * 1024 * 1024  # 1 GB

db = SQLAlchemy(app)
migrate = Migrate(app, db)






class ClientKYC(db.Model):
    __tablename__ = "client_kyc"

    id = db.Column(db.Integer, primary_key=True)
    client_code = db.Column(db.String(20), unique=True, index=True)

    # BUSINESS
    company_name = db.Column(db.String(200))
    gst_number = db.Column(db.String(30))
    company_pan = db.Column(db.String(20))
    msme_registration = db.Column(db.String(50))
    iec_code = db.Column(db.String(30))
    address = db.Column(db.Text)

    # BANK
    bank_name = db.Column(db.String(150))
    account_number = db.Column(db.String(50))
    ifsc_code = db.Column(db.String(20))
    branch = db.Column(db.String(150))

    # DOCUMENTS
    pan_doc = db.Column(LONGBLOB)
    gst_doc = db.Column(LONGBLOB)
    msme_doc = db.Column(LONGBLOB)
    iec_doc = db.Column(LONGBLOB)
    visiting_card_doc = db.Column(LONGBLOB)
    bank_doc = db.Column(LONGBLOB)
    
class ClientOwner(db.Model):
    __tablename__ = "client_owner"

    id = db.Column(db.Integer, primary_key=True)

    client_id = db.Column(db.Integer, db.ForeignKey("client_kyc.id"))

    full_name = db.Column(db.String(150))
    pan_number = db.Column(db.String(20))
    aadhar_number = db.Column(db.String(20))
    phone_number = db.Column(db.String(20))
    email_id = db.Column(db.String(150))
    address = db.Column(db.Text)

    pan_doc = db.Column(LONGBLOB)
    aadhar_doc = db.Column(LONGBLOB)
    
class ClientMOU(db.Model):
    __tablename__ = "client_mou"

    id = db.Column(db.Integer, primary_key=True)

    client_id = db.Column(db.Integer, db.ForeignKey("client_kyc.id"), nullable=False)

    client = db.relationship("ClientKYC", backref="mou")

    gold_credit_days = db.Column(db.Integer)
    labour_credit_days = db.Column(db.Integer)
    diamond_credit_days = db.Column(db.Integer)
    chain_credit_days = db.Column(db.Integer)
    certificate_hallmark_days = db.Column(db.Integer)

    making_charges = db.Column(db.Numeric(10,2))
    chain_charges = db.Column(db.Numeric(10,2))
    synthetic_stone_charges = db.Column(db.Numeric(10,2))

    minimum_labour_rule = db.Column(db.String(100))
    nose_pin_labour = db.Column(db.String(50))

    gold_labour_percent = db.Column(db.Numeric(5,2))

    # ⭐ NEW FIELDS
    diamond_clarity = db.Column(db.String(20))
    diamond_color = db.Column(db.String(20))

    diamond_rate_2 = db.Column(db.Numeric(10,2))
    diamond_rate_6_5 = db.Column(db.Numeric(10,2))
    diamond_rate_11 = db.Column(db.Numeric(10,2))

    mou_document = db.Column(LONGBLOB)
    
class ProductionOrder(db.Model):
    __tablename__ = "production_orders"

    id = db.Column(db.Integer, primary_key=True)
    
    order_no = db.Column(db.String(20), unique=True, nullable=False)

    client_id = db.Column(db.Integer, db.ForeignKey("client_kyc.id"), nullable=False)
    client = db.relationship("ClientKYC", backref="production_orders")

    order_datetime = db.Column(db.Date, nullable=False)
    delivery_datetime = db.Column(db.Date)

    total_amount = db.Column(db.Float, default=0)
    remark = db.Column(db.Text)

    items = db.relationship(
        "ProductionOrderItem",
        backref="order",
        cascade="all, delete-orphan"
    )


class ProductionOrderItem(db.Model):
    __tablename__ = "production_order_items"

    id = db.Column(db.Integer, primary_key=True)

    order_id = db.Column(
        db.Integer,
        db.ForeignKey("production_orders.id"),
        nullable=False
    )

    style_no = db.Column(db.String(100), nullable=False)

    diamond_clarity = db.Column(db.String(20))
    gold_color = db.Column(db.String(30))
    diamond_color = db.Column(db.String(20))

    gold_purity = db.Column(db.String(20))
    gold_purity_factor = db.Column(db.Float)

    pieces = db.Column(db.Integer, default=1)
    style_remark = db.Column(db.Text)


class StyleNo(db.Model):
    __tablename__ = "style_no"

    id = db.Column(db.Integer, primary_key=True)
    style_no = db.Column(db.String(100), nullable=False, index=True)
    
    category = db.Column(db.String(100))

    gold_wt = db.Column(db.Float)
    net_wt = db.Column(db.Float)
    dia_wt = db.Column(db.Float)
    dia_pc = db.Column(db.Integer)
        # New Columns
    cstn_pc = db.Column(db.Integer)
    cstn_wt = db.Column(db.Float)

    brand = db.Column(db.String(100))

    image = db.Column(LONGBLOB)
    gem_chart = db.Column(LONGBLOB)


class DiamondSieveMaster(db.Model):
    __tablename__ = "diamond_sieve_master"

    id = db.Column(db.Integer, primary_key=True)

    sieve_range = db.Column(db.String(50), nullable=False)
    sieve_size = db.Column(db.String(20), nullable=False)
    mm_size = db.Column(db.Float, nullable=False)
    no_of_stones = db.Column(db.Integer, nullable=False)
    ct_weight_per_piece = db.Column(db.Float, nullable=False)



class User(db.Model):
    __tablename__ = "users"

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(50), nullable=False)

class ProductionStatus(db.Model):
    __tablename__ = "production_status"

    id = db.Column(db.Integer, primary_key=True)

    order_id = db.Column(db.Integer, db.ForeignKey("production_orders.id"), unique=True, nullable=False)
    order = db.relationship("ProductionOrder", backref=db.backref("status", uselist=False))
    
    batch_no = db.Column(db.String(50), nullable=True)

    # stores: casting, pre_filing, pre_ep, ... final_qc
    data = db.Column(JSON, nullable=False, default=dict)

    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    
class GoldStock(db.Model):
    __tablename__ = "gold_stock"

    id = db.Column(db.Integer, primary_key=True)

    # ✅ Identify stock type
    stock_type = db.Column(db.String(20), nullable=False)  
    # values: "gold", "silver", "alloy"

    # ✅ Gold fields
    gold_type = db.Column(db.String(30))   # Yellow, Rose, White
    purity = db.Column(db.String(10))      # 9K, 14K, 18K, 22K

    weight = db.Column(db.Float, default=0)   # Gold weight
    pieces = db.Column(db.Integer, default=0) # pieces

    # ✅ Silver & Alloy
    silver_weight = db.Column(db.Float, default=0)
    alloy_weight = db.Column(db.Float, default=0)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
class DiamondStock(db.Model):

    __tablename__ = "diamond_stock"

    id = db.Column(db.Integer, primary_key=True)
    size = db.Column(db.String(20))
    clarity = db.Column(db.String(10))
    color = db.Column(db.String(10))
    pieces = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
class Employee(db.Model):
    __tablename__ = "employees"

    id = db.Column(db.Integer, primary_key=True)

    emp_code = db.Column(db.String(20), unique=True)
    name = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    role = db.Column(db.String(50))
    
class WaxIssueOrder(db.Model):
    __tablename__ = "wax_issue_orders"

    id = db.Column(db.Integer, primary_key=True)

    issue_id = db.Column(
        db.Integer,
        db.ForeignKey("wax_issues.id"),
        nullable=False
    )

    order_id = db.Column(
        db.Integer,
        db.ForeignKey("production_orders.id"),
        nullable=False
    )
    
class WaxIssue(db.Model):
    __tablename__ = "wax_issues"

    id = db.Column(db.Integer, primary_key=True)

    issue_code = db.Column(db.String(20), unique=True)

    wax_wt = db.Column(db.Float)
    total_gold_wt = db.Column(db.Float)

    silver_wt = db.Column(db.Float)
    alloy_wt = db.Column(db.Float)
    final_total_wt = db.Column(db.Float)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
class WaxIssueStyle(db.Model):
    __tablename__ = "wax_issue_styles"

    id = db.Column(db.Integer, primary_key=True)

    issue_id = db.Column(
        db.Integer,
        db.ForeignKey("wax_issues.id"),
        nullable=False
    )

    style_no = db.Column(db.String(100))

    gold_color = db.Column(db.String(30))
    gold_purity = db.Column(db.String(20))
    order_id = db.Column(
        db.Integer,
        db.ForeignKey("production_orders.id")
    )
    
class WaxIssueGold(db.Model):
    __tablename__ = "wax_issue_gold"

    id = db.Column(db.Integer, primary_key=True)

    issue_id = db.Column(
        db.Integer,
        db.ForeignKey("wax_issues.id"),
        nullable=False
    )

    gold_gm = db.Column(db.Float)
    gold_kt = db.Column(db.String(10))

    # ✅ NEW FIELDS
    gold_color = db.Column(db.String(30))
    pieces = db.Column(db.Integer)
    
class DiamondIssue(db.Model):
    __tablename__ = "diamond_issues"

    id = db.Column(db.Integer, primary_key=True)

    issue_code = db.Column(db.String(20), unique=True)

    total_pcs = db.Column(db.Integer)
    total_ct = db.Column(db.Float)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
class DiamondIssueOrder(db.Model):
    __tablename__ = "diamond_issue_orders"

    id = db.Column(db.Integer, primary_key=True)

    issue_id = db.Column(
        db.Integer,
        db.ForeignKey("diamond_issues.id"),
        nullable=False
    )

    order_id = db.Column(
        db.Integer,
        db.ForeignKey("production_orders.id"),
        nullable=False
    )
    
class DiamondIssueStyle(db.Model):
    __tablename__ = "diamond_issue_styles"

    id = db.Column(db.Integer, primary_key=True)

    issue_id = db.Column(
        db.Integer,
        db.ForeignKey("diamond_issues.id"),
        nullable=False
    )

    style_no = db.Column(db.String(100))

    clarity = db.Column(db.String(20))
    color = db.Column(db.String(20))
    pcs = db.Column(db.Integer)
    
class DiamondIssueRound(db.Model):
    __tablename__ = "diamond_issue_rounds"

    id = db.Column(db.Integer, primary_key=True)

    issue_id = db.Column(
        db.Integer,
        db.ForeignKey("diamond_issues.id"),
        nullable=False
    )

    size = db.Column(db.String(20))
    pcs = db.Column(db.Integer)
    ct = db.Column(db.Float)
    
    

 

# -----------------------------
# MAIN ROUTES
# -----------------------------

# ======================================
# 🔧 GETTING DIAMOND DETAILS FROM DB
# ======================================
@app.route("/contact", methods=["POST"])
def contact():

    data = request.get_json()

    name = data.get("name")
    email = data.get("email")
    subject = data.get("subject")
    message = data.get("message")

    msg = Message(
        subject=f"Website Contact - {subject}",
        recipients=["sonalikajewellers2021@gmail.com"]
    )

    msg.body = f"""
Name : {name}

Email : {email}

Subject : {subject}

Message :
{message}
"""

    try:

        mail.send(msg)

        return jsonify({
            "success": True,
            "message": "Message sent successfully."
        })

    except Exception as e:

        print(e)

        return jsonify({
            "success": False,
            "message": "Failed to send message."
        }),500


@app.route("/ledger")
def ledger():
    return render_template("ledger.html")

@app.route("/sales_er")
def sales_er():
    return render_template("sales_er.html")

@app.route("/sales_nk")
def sales_nk():
    return render_template("sales_nk.html")

@app.route("/sales_ns")
def sales_ns():
    return render_template("sales_ns.html")


@app.route("/sales_lr")
def sales_lr():
    return render_template("sales_lr.html")

@app.route("/sales_pd")
def sales_pd():
    return render_template("sales_pd.html")

@app.route("/sales_tbrl")
def sales_tbrl():
    return render_template("sales_tbrl.html")

@app.route("/sales_ter")
def sales_ter():
    return render_template("sales_ter.html")

@app.route("/sales_tlr")
def sales_tlr():
    return render_template("sales_tlr.html")

@app.route("/sales_tnk")
def sales_tnk():
    return render_template("sales_tnk.html")

@app.route("/sales_tpd")
def sales_tpd():
    return render_template("sales_tpd.html")


@app.route("/api/styles", methods=["GET"])
def api_styles():

    brand = request.args.get("brand", "").strip()
    category = request.args.get("category", "").strip()
    page = int(request.args.get("page", 1))

    PER_PAGE = 8

    query = StyleNo.query

    # Filter by Brand
    if brand:
        query = query.filter(StyleNo.brand.ilike(brand))

    # Filter by Category
    if category:
        query = query.filter(StyleNo.category.ilike(category))

    total = query.count()

    styles = (
        query
        .order_by(StyleNo.id)
        .offset((page - 1) * PER_PAGE)
        .limit(PER_PAGE)
        .all()
    )

    products = []

    for style in styles:

        products.append({
            "id": style.id,
            "style_no": style.style_no,
            "category": style.category,
            "gold_wt": style.gold_wt,
            "net_wt": style.net_wt,
            "dia_wt": style.dia_wt,
            "dia_pc": style.dia_pc,
            "brand": style.brand,
            "cstn_pc": style.cstn_pc,
            "cstn_wt":style.cstn_wt,
            "image": base64.b64encode(style.image).decode("utf-8") if style.image else ""
        })

    return jsonify({
        "page": page,
        "per_page": PER_PAGE,
        "total": total,
        "products": products
    })

@app.route("/jobcard_list")
def jobcard_list():
    return render_template("jobcard_list.html")

@app.route("/get-job-card/<int:issue_id>/<style_no>")
def get_job_card(issue_id, style_no):

    try:
        style_no = (style_no or "").strip()

        # ===============================
        # GET ALL STYLES OF ISSUE
        # ===============================
        styles = WaxIssueStyle.query.filter_by(issue_id=issue_id).all()

        # SORT (important for consistent numbering)
        styles_sorted = sorted(styles, key=lambda x: x.style_no)

        # ===============================
        # FIND INDEX
        # ===============================
        index = 1
        for i, s in enumerate(styles_sorted, start=1):
            if (s.style_no or "").strip() == style_no:
                index = i
                break

        # ===============================
        # 🔥 GENERATE JC NUMBER
        # ===============================
        job_card_no = f"JC{str(index).zfill(3)}"

        # ===============================
        # GET STYLE
        # ===============================
        style = next(
            (s for s in styles if (s.style_no or "").strip() == style_no),
            None
        )

        if not style:
            return jsonify({"error": "Style not found"}), 404

        # ===============================
        # ORDER + CLIENT
        # ===============================
        order_no = "-"
        client_code = "-"

        if style.order_id:
            order = db.session.get(ProductionOrder, style.order_id)

            if order:
                order_no = order.order_no

                if order.client_id:
                    client = db.session.get(ClientKYC, order.client_id)
                    if client:
                        client_code = client.client_code

        # ===============================
        # DIAMOND
        # ===============================
        diamonds = DiamondIssueStyle.query.filter(
            db.func.trim(DiamondIssueStyle.style_no) == style_no
        ).all()

        diamond_quality = ", ".join({
            f"{(d.clarity or '-')} | {(d.color or '-')}"
            for d in diamonds
        }) or "-"

        return jsonify({
            "job_card_no": job_card_no,
            "order_no": order_no,
            "client_code": client_code,
            "style_no": style.style_no,
            "gold_color": style.gold_color,
            "gold_purity": style.gold_purity,
            "diamond_quality": diamond_quality
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/print-wax-issue/<int:id>")
def print_wax_issue(id):

    w = WaxIssue.query.get_or_404(id)

    order_links = WaxIssueOrder.query.filter_by(issue_id=w.id).all()

    order_nos = []
    total_pcs = 0

    for link in order_links:
        order = ProductionOrder.query.get(link.order_id)
        if not order:
            continue

        order_nos.append(order.order_no)

        for item in order.items:
            total_pcs += item.pieces or 0

    styles = WaxIssueStyle.query.filter_by(issue_id=w.id).all()
    gold = WaxIssueGold.query.filter_by(issue_id=w.id).all()

    return render_template(
        "print_wax_issue.html",
        issue=w,
        orders=order_nos,
        pieces=total_pcs,
        styles=styles,
        gold=gold
    )

@app.route("/wax-issues")
def wax_issues_page():
    return render_template("wax_issues_list.html")

@app.route("/mou-print")
def mou_print():
    return render_template("mou_print.html")

@app.route("/print-kyc")
def print_kyc():
    return render_template("print_kyc.html")

@app.route("/print-order/<int:order_id>")
def print_page(order_id):
    return render_template("print_order.html", order_id=order_id)

def size_to_mm(size_text):
    m = re.search(r"(\d+(\.\d+)?)", str(size_text))
    return float(m.group(1)) if m else None


@app.route("/api/order/<int:order_id>")
def get_order_api(order_id):

    order = ProductionOrder.query\
        .options(
            db.joinedload(ProductionOrder.client),
            db.joinedload(ProductionOrder.items)
        ).get_or_404(order_id)

    items_data = []

    for i in order.items:

        # 🔹 ORIGINAL FIELDS (KEEP SAME)
        style_no = i.style_no
        diamond_clarity = i.diamond_clarity
        gold_color = i.gold_color
        diamond_color = i.diamond_color
        gold_purity = i.gold_purity
        qty = i.pieces or 1
        remark = i.style_remark

        # 🔹 GET STYLE DATA
        style_rows = StyleNo.query.filter_by(style_no=style_no).all()

        gross_wt = 0
        diamond_pcs = 0
        diamond_wt = 0

        for r in style_rows:

            # GOLD WT (take first available)
            if r.gold_wt and gross_wt == 0:
                gross_wt = float(r.gold_wt)

            # DIAMOND PCS
            if r.pcs:
                diamond_pcs += int(r.pcs)

            # MM CONVERSION
            mm = size_to_mm(r.size)

            if mm:
                sieve = (
                    DiamondSieveMaster.query
                    .filter(DiamondSieveMaster.mm_size <= mm)
                    .order_by(
                        DiamondSieveMaster.mm_size.desc(),
                        DiamondSieveMaster.id.asc()
                    )
                    .first()
                )

                if sieve and r.pcs:
                    diamond_wt += int(r.pcs) * float(sieve.ct_weight_per_piece)

        # 🔹 APPLY QUANTITY
        gross_wt *= qty
        diamond_pcs *= qty
        diamond_wt *= qty

        # 🔹 NET WT (same logic as frontend)
        net_wt = gross_wt - (diamond_wt * 0.2)

        # 🔹 FINAL ITEM DATA (MERGED)
        items_data.append({
            # ✅ OLD FIELDS
            "style_no": style_no,
            "diamond_clarity": diamond_clarity,
            "gold_color": gold_color,
            "diamond_color": diamond_color,
            "gold_purity": gold_purity,
            "pieces": qty,
            "remark": remark,

            # ✅ NEW FIELDS (ADDED)
            "gross_weight": round(gross_wt, 3),
            "net_weight": round(net_wt, 3),
            "diamond_weight": round(diamond_wt, 3),
            "diamond_pcs": diamond_pcs
        })

    return jsonify({
        # ✅ ORIGINAL RESPONSE (UNCHANGED)
        "order_id": order.id,
        "order_no": order.order_no,

        "order_date": str(order.order_datetime),
        "delivery_date": str(order.delivery_datetime),

        "total_amount": order.total_amount,
        "remark": order.remark,

        "company": {
            "name": "SONALIKA JEWELLERS",
            "gst": "29BQJPS8449F1Z3",
            "address": "5 GOODS SHED ROAD CHELUVADIPALYA BANGALORE 5600"
        },

        "client": {
            "client_code": order.client.client_code,
            "company_name": order.client.company_name,
            "gst": order.client.gst_number,
            "address": order.client.address
        },

        # ✅ MERGED ITEMS
        "items": items_data
    })

@app.route("/api/wax-issues-full", methods=["GET"])
def get_wax_issues_full():

    issues = WaxIssue.query.order_by(WaxIssue.id.desc()).all()
    result = []

    for w in issues:

        # =========================
        # ORDERS UNDER THIS ISSUE
        # =========================
        order_links = WaxIssueOrder.query.filter_by(issue_id=w.id).all()

        order_nos = []
        total_pcs = 0

        for link in order_links:
            order = ProductionOrder.query.get(link.order_id)

            if not order:
                continue

            order_nos.append(order.order_no)

            for item in order.items:
                total_pcs += item.pieces or 0

        # =========================
        # BATCH NO (IMPORTANT FIX)
        # =========================
        batch_no = None

        for link in order_links:
            status = ProductionStatus.query.filter_by(order_id=link.order_id).first()

            if status and status.batch_no:
                batch_no = status.batch_no
                break

        # =========================
        # STYLE DETAILS
        # =========================
        style_rows = WaxIssueStyle.query.filter_by(issue_id=w.id).all()

        styles = []
        for s in style_rows:
            styles.append({
                "style_no": s.style_no,
                "gold_color": s.gold_color,
                "gold_purity": s.gold_purity
            })

        # =========================
        # GOLD DATA
        # =========================
        gold_rows = WaxIssueGold.query.filter_by(issue_id=w.id).all()

        gold_data = []
        for g in gold_rows:
            gold_data.append({
                "gold_gm": g.gold_gm,
                "gold_kt": g.gold_kt
            })

        # =========================
        # DIAMOND DATA
        # =========================
        diamond_summary = []
        total_diamond_ct = 0
        processed_issue_ids = set()

        for link in order_links:

            d_links = DiamondIssueOrder.query.filter_by(order_id=link.order_id).all()

            for dlink in d_links:

                if dlink.issue_id in processed_issue_ids:
                    continue

                processed_issue_ids.add(dlink.issue_id)

                issue = DiamondIssue.query.get(dlink.issue_id)
                if not issue:
                    continue

                total_diamond_ct += issue.total_ct or 0

                # STYLE DIAMONDS
                styles_d = DiamondIssueStyle.query.filter_by(issue_id=issue.id).all()
                for s in styles_d:
                    diamond_summary.append(
                        f"{s.style_no} {s.clarity} | {s.color} | {s.pcs} pcs"
                    )

                # ROUND DIAMONDS
                rounds = DiamondIssueRound.query.filter_by(issue_id=issue.id).all()
                for r in rounds:
                    diamond_summary.append(
                        f"{r.size} {r.pcs} pcs | {r.ct} ct"
                    )

        # =========================
        # FINAL RESPONSE
        # =========================
        result.append({
            "id": w.id,
            "order_no": ", ".join(order_nos),
            "batch_no": batch_no,
            "date": w.created_at.strftime("%d-%m-%Y"),
            "styles": styles,
            "pieces": total_pcs,
            "gold": gold_data,
            "total_gold_wt": w.total_gold_wt or 0,
            "diamond_text": diamond_summary,
            "total_diamond_ct": total_diamond_ct
        })

    return jsonify({"data": result})
# ======================================
# 🔧 HELPER: NORMALIZE SIZE
# ======================================
def normalize_size(size_str):
    nums = re.findall(r"\d+(?:\.\d+)?", size_str)

    if len(nums) >= 2:
        return f"{float(nums[0])}*{float(nums[1])}"

    return size_str.replace("x", "*")


# ======================================
# 🚀 API: CREATE DIAMOND ISSUE
# ======================================

@app.route("/api/diamond-issue", methods=["POST"])
def create_diamond_issue():

    data = request.get_json()

    try:
        summary = data.get("summary", [])
        rounds = data.get("rounds", [])

        # =====================================
        # 🔥 STEP 1: BUILD REQUIRED MAP (CORRECT)
        # =====================================
        required_map = {}

        for r in rounds:
            size = normalize_size(r["size"])
            pcs = int(r["pcs"] or 0)

            for s in summary:
                clarity = s["clarity"]
                color = s["color"]

                key = (size, clarity, color)
                required_map[key] = required_map.get(key, 0) + pcs

        # =====================================
        # 🔥 STEP 2: CHECK STOCK
        # =====================================
        errors = []

        for (size, clarity, color), req_pcs in required_map.items():

            stock = db.session.execute(text("""
                SELECT SUM(pieces) as total
                FROM diamond_stock
                WHERE size = :size
                AND clarity = :clarity
                AND color = :color
            """), {
                "size": size,
                "clarity": clarity,
                "color": color
            }).fetchone()

            available = stock.total or 0

            if available < req_pcs:
                errors.append(
                    f"{size} | {clarity} | {color} → Need {req_pcs}, Available {available}"
                )

        # ❌ STOP IF ERROR
        if errors:
            return jsonify({
                "ok": False,
                "message": "Stock not available",
                "details": errors
            }), 400

        # =====================================
        # ✅ STEP 3: SAVE ISSUE
        # =====================================
        issue = DiamondIssue(
            issue_code=f"DI{int(time.time())}",
            total_pcs=data.get("total_pcs"),
            total_ct=data.get("total_ct")
        )
        db.session.add(issue)
        db.session.flush()

        # Orders
        for oid in data.get("orders", []):
            db.session.add(DiamondIssueOrder(
                issue_id=issue.id,
                order_id=oid
            ))

        # Summary
        for s in summary:
            db.session.add(DiamondIssueStyle(
                issue_id=issue.id,
                style_no=s["style_no"],
                clarity=s["clarity"],
                color=s["color"],
                pcs=s["pcs"]
            ))

        # Rounds
        for r in rounds:
            db.session.add(DiamondIssueRound(
                issue_id=issue.id,
                size=normalize_size(r["size"]),
                pcs=r["pcs"],
                ct=r["ct"]
            ))

        # =====================================
        # 🔥 STEP 4: DEDUCT STOCK (CORRECT)
        # =====================================
        for (size, clarity, color), req_pcs in required_map.items():

            rows = db.session.execute(text("""
                SELECT id, pieces
                FROM diamond_stock
                WHERE size = :size
                AND clarity = :clarity
                AND color = :color
                ORDER BY id
            """), {
                "size": size,
                "clarity": clarity,
                "color": color
            }).fetchall()

            remaining = req_pcs

            for row in rows:

                if remaining <= 0:
                    break

                deduct = min(row.pieces, remaining)

                db.session.execute(text("""
                    UPDATE diamond_stock
                    SET pieces = pieces - :deduct
                    WHERE id = :id
                """), {
                    "deduct": deduct,
                    "id": row.id
                })

                remaining -= deduct

        db.session.commit()

        return jsonify({"ok": True})

    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "message": str(e)}), 500

@app.route("/api/multi-order-styles", methods=["POST"])
def multi_order_styles():

    try:
        data = request.get_json()
        order_ids = list(map(int, data.get("order_ids", [])))

        styles = []

        for oid in order_ids:

            items = ProductionOrderItem.query.filter_by(order_id=oid).all()

            for s in items:
                styles.append({
                    "style_no": s.style_no,
                    "pieces": s.pieces,
                    "gold_color": s.gold_color,
                    "gold_purity": s.gold_purity,
                    "diamond_clarity": s.diamond_clarity,
                    "diamond_color": s.diamond_color,
                    "remark": s.style_remark   # ✅ correct field name
                })

        return jsonify({"ok": True, "styles": styles})

    except Exception as e:
        print("ERROR:", str(e))
        return jsonify({"ok": False, "message": str(e)}), 500

@app.route("/api/wax-issues", methods=["GET"])
def get_wax_issues():

    issues = WaxIssue.query.order_by(WaxIssue.id.desc()).all()

    data = []

    for i, issue in enumerate(issues, start=1):

        # ✅ STYLES
        styles = WaxIssueStyle.query.filter_by(issue_id=issue.id).all()
        style_list = [{
            "style_no": s.style_no,
            "gold_color": s.gold_color,
            "gold_purity": s.gold_purity
        } for s in styles]

        # ✅ GOLD ROWS
        golds = WaxIssueGold.query.filter_by(issue_id=issue.id).all()
        gold_list = [{
            "gold_gm": g.gold_gm,
            "gold_kt": g.gold_kt
        } for g in golds]

        data.append({
            "sl_no": i,
            "id": issue.id,
            "order_no": issue.issue_code,
            "styles": style_list,
            "gold": gold_list,
            "no_of_pieces": len(style_list),

            # ✅ WEIGHTS
            "wax_wt": issue.wax_wt,
            "total_gold_wt": issue.total_gold_wt,
            "silver_wt": issue.silver_wt,
            "alloy_wt": issue.alloy_wt,
            "final_total_wt": issue.final_total_wt,

            "date": issue.created_at.strftime("%d-%m-%Y")
        })

    return jsonify({"status": "success", "data": data})

@app.route("/submit-wax-issue", methods=["POST"])
def submit_wax_issue():

    data = request.get_json()

    # ===============================
    # SAFE CAST
    # ===============================
    def to_float(val):
        try:
            return float(val)
        except:
            return 0

    def to_int(val):
        try:
            return int(val)
        except:
            return 0

    # ===============================
    # GET DATA
    # ===============================
    order_ids = data.get("order_ids", [])   # ✅ NEW (array)
    order_no = data.get("order_no")

    wax_wt = to_float(data.get("wax_wt"))
    total_gold_wt = to_float(data.get("total_gold_wt"))
    silver_wt = to_float(data.get("silver_wt"))
    alloy_wt = to_float(data.get("alloy_wt"))
    final_total_wt = to_float(data.get("final_total_wt"))

    styles = data.get("styles", [])
    gold_rows = data.get("gold_rows", [])

    # ===============================
    # VALIDATION
    # ===============================
    if not order_ids or not isinstance(order_ids, list):
        return jsonify({"status": "error", "message": "Order IDs required"})

    if not order_no:
        return jsonify({"status": "error", "message": "Order No missing"})

    if not gold_rows:
        return jsonify({"status": "error", "message": "Gold data required"})

    try:
        from collections import defaultdict

        # ===============================
        # 1. GROUP GOLD (KT + COLOR)
        # ===============================
        stock_map = defaultdict(lambda: {"gm": 0, "pcs": 0})

        for g in gold_rows:
            gm = to_float(g.get("gold_gm"))
            kt = g.get("gold_kt")
            color = g.get("gold_color")
            pcs = to_int(g.get("pieces"))

            if gm > 0 and kt and color:
                kt_db = f"{kt}K"
                stock_map[(kt_db, color)]["gm"] += gm
                stock_map[(kt_db, color)]["pcs"] += pcs

        # ===============================
        # 2. CHECK GOLD STOCK (GM + PCS)
        # ===============================
        for (kt, color), val in stock_map.items():

            total_gm = val["gm"]
            total_pcs = val["pcs"]

            available_gm = db.session.query(func.sum(GoldStock.weight))\
                .filter(
                    GoldStock.stock_type == "gold",
                    GoldStock.purity == kt,
                    GoldStock.gold_type == color
                ).scalar() or 0

            available_pcs = db.session.query(func.sum(GoldStock.pieces))\
                .filter(
                    GoldStock.stock_type == "gold",
                    GoldStock.purity == kt,
                    GoldStock.gold_type == color
                ).scalar() or 0

            if available_gm < total_gm:
                return jsonify({
                    "status": "no_stock",
                    "message": f"Not enough {color} {kt} gold (gm)"
                })

            if available_pcs < total_pcs:
                return jsonify({
                    "status": "no_stock",
                    "message": f"Not enough {color} {kt} pieces"
                })

        # ===============================
        # 3. CHECK SILVER
        # ===============================
        if silver_wt > 0:
            silver_available = db.session.query(func.sum(GoldStock.silver_weight))\
                .filter(GoldStock.stock_type == "silver").scalar() or 0

            if silver_available < silver_wt:
                return jsonify({
                    "status": "no_stock",
                    "message": "Not enough Silver stock"
                })

        # ===============================
        # 4. CHECK ALLOY
        # ===============================
        if alloy_wt > 0:
            alloy_available = db.session.query(func.sum(GoldStock.alloy_weight))\
                .filter(GoldStock.stock_type == "alloy").scalar() or 0

            if alloy_available < alloy_wt:
                return jsonify({
                    "status": "no_stock",
                    "message": "Not enough Alloy stock"
                })

        # ===============================
        # 5. SAVE ISSUE (NO order_id here)
        # ===============================
        issue = WaxIssue(
            issue_code=order_no,
            wax_wt=wax_wt,
            total_gold_wt=total_gold_wt,
            silver_wt=silver_wt,
            alloy_wt=alloy_wt,
            final_total_wt=final_total_wt
        )

        db.session.add(issue)
        db.session.flush()

        # ===============================
        # 6. SAVE MULTIPLE ORDERS ✅ NEW
        # ===============================
        for oid in order_ids:
            db.session.add(WaxIssueOrder(
                issue_id=issue.id,
                order_id=oid
            ))

        # ===============================
        # 7. SAVE STYLES
        # ===============================
        # ✅ FIXED VERSION
        for s in styles:
        
            style_no = s.get("style_no")
        
            # 🔥 FIND ORDER FROM production_order_items
            item = ProductionOrderItem.query.filter_by(
                style_no=style_no
            ).first()
        
            order_id = item.order_id if item else None
        
            db.session.add(WaxIssueStyle(
                issue_id=issue.id,
                style_no=style_no,
                gold_color=s.get("gold_color"),
                gold_purity=s.get("gold_purity"),
                order_id=order_id   # ✅ AUTO SET
            ))
        
        # ===============================
        # 8. SAVE GOLD ROWS
        # ===============================
        for g in gold_rows:

            gm = to_float(g.get("gold_gm"))
            kt = g.get("gold_kt")
            color = g.get("gold_color")
            pcs = to_int(g.get("pieces"))

            if gm > 0 and kt and color:
                db.session.add(WaxIssueGold(
                    issue_id=issue.id,
                    gold_gm=gm,
                    gold_kt=f"{kt}K",
                    gold_color=color,
                    pieces=pcs
                ))

        # ===============================
        # 9. DEDUCT GOLD (FIFO + GM + PCS)
        # ===============================
        for (kt, color), val in stock_map.items():

            remaining_gm = val["gm"]
            remaining_pcs = val["pcs"]

            stocks = GoldStock.query.filter_by(
                stock_type="gold",
                purity=kt,
                gold_type=color
            ).order_by(GoldStock.id.asc()).all()

            for stock in stocks:

                if remaining_gm <= 0 or remaining_pcs <= 0:
                    break

                if stock.weight <= 0 or stock.pieces <= 0:
                    continue

                per_piece_weight = stock.weight / stock.pieces

                pcs_take = min(stock.pieces, remaining_pcs)
                weight_take = pcs_take * per_piece_weight

                if weight_take > remaining_gm:
                    weight_take = remaining_gm
                    pcs_take = int(weight_take / per_piece_weight)

                stock.weight -= weight_take
                stock.pieces -= pcs_take

                remaining_gm -= weight_take
                remaining_pcs -= pcs_take

        # ===============================
        # 10. DEDUCT SILVER
        # ===============================
        if silver_wt > 0:
            remaining = silver_wt
            stocks = GoldStock.query.filter_by(stock_type="silver").all()

            for s in stocks:
                if remaining <= 0:
                    break

                if s.silver_weight >= remaining:
                    s.silver_weight -= remaining
                    remaining = 0
                else:
                    remaining -= s.silver_weight
                    s.silver_weight = 0

        # ===============================
        # 11. DEDUCT ALLOY
        # ===============================
        if alloy_wt > 0:
            remaining = alloy_wt
            stocks = GoldStock.query.filter_by(stock_type="alloy").all()

            for a in stocks:
                if remaining <= 0:
                    break

                if a.alloy_weight >= remaining:
                    a.alloy_weight -= remaining
                    remaining = 0
                else:
                    remaining -= a.alloy_weight
                    a.alloy_weight = 0

        # ===============================
        # COMMIT
        # ===============================
        db.session.commit()

        return jsonify({
            "status": "success",
            "issue_code": order_no
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            "status": "error",
            "message": str(e)
        })
    
@app.route("/api/employees/<int:id>", methods=["DELETE"])
def delete_employee(id):

    emp = Employee.query.get(id)

    if not emp:
        return jsonify({"ok": False})

    db.session.delete(emp)
    db.session.commit()

    return jsonify({"ok": True})

@app.route("/api/employees")
def get_employees():

    employees = Employee.query.order_by(Employee.id.desc()).all()

    return jsonify({
        "ok": True,
        "employees": [
            {
                "id": e.id,
                "emp_code": e.emp_code,
                "name": e.name,
                "phone": e.phone,
                "role": e.role
            }
            for e in employees
        ]
    })

@app.route("/api/employees", methods=["POST"])
def create_employees():

    data = request.get_json()
    employees = data.get("employees", [])

    # get last employee
    last = Employee.query.order_by(Employee.id.desc()).first()

    if last and last.emp_code:
        last_num = int(last.emp_code.replace("SJED", ""))
    else:
        last_num = 0

    new_list = []

    for i, emp in enumerate(employees, start=1):

        new_code = f"SJED{str(last_num + i).zfill(4)}"

        new_emp = Employee(
            emp_code=new_code,
            name=emp.get("name"),
            phone=emp.get("phone"),
            role=emp.get("role")
        )

        db.session.add(new_emp)
        new_list.append(new_code)

    db.session.commit()

    return jsonify({
        "ok": True,
        "codes": new_list
    })

@app.route("/add-diamond", methods=["POST"])
def add_diamond():

    data = request.get_json()

    size = data.get("size")
    clarity = data.get("clarity")
    color = data.get("color")
    pieces = data.get("pieces")

    new_diamond = DiamondStock(
        size=size,
        clarity=clarity,
        color=color,
        pieces=pieces
    )

    db.session.add(new_diamond)
    db.session.commit()

    return jsonify({
        "status": "success",
        "message": "Diamond added successfully"
    })
    
@app.route("/get-diamond-stock")
def get_diamond_stock():

    page = request.args.get("page", 1, type=int)
    per_page = 7

    diamonds = DiamondStock.query.order_by(
        DiamondStock.id.desc()
    ).paginate(page=page, per_page=per_page)

    data = []

    for d in diamonds.items:

        data.append({
            "id": d.id,
            "size": d.size,
            "clarity": d.clarity,
            "color": d.color,
            "pieces": d.pieces,
            "date": d.created_at.strftime("%d-%m-%Y")
        })

    return jsonify({
        "data": data,
        "page": page,
        "total_pages": diamonds.pages
    })

@app.route("/add-metal", methods=["POST"])
def add_metal():

    data = request.get_json()

    metal_type = data.get("metal_type")  # silver / alloy
    weight = float(data.get("weight", 0))

    metal = GoldStock(
        stock_type=metal_type,   # ✅ ADD THIS
        silver_weight=weight if metal_type == "silver" else 0,
        alloy_weight=weight if metal_type == "alloy" else 0
    )

    db.session.add(metal)
    db.session.commit()

    return jsonify({"status": "success"})

@app.route("/get-metal-stock")
def get_metal_stock():

    page = request.args.get("page", 1, type=int)
    per_page = 5

    stocks = GoldStock.query.filter(
        GoldStock.stock_type.in_(["silver", "alloy"])
    ).order_by(GoldStock.id.desc()).paginate(page=page, per_page=per_page)

    data = []

    for s in stocks.items:
    
        if s.stock_type == "silver":
            metal_type = "Silver"
            weight = s.silver_weight
        else:
            metal_type = "Alloy"
            weight = s.alloy_weight

        data.append({
            "id": s.id,
            "metal_type": metal_type,
            "weight": round(weight, 3),
            "date": s.created_at.strftime("%d-%m-%Y")
        })

    # ✅ Calculate totals
    total_silver = db.session.query(func.sum(GoldStock.silver_weight)).scalar() or 0
    total_alloy = db.session.query(func.sum(GoldStock.alloy_weight)).scalar() or 0

    return jsonify({
        "data": data,
        "page": page,
        "total_pages": stocks.pages,
        "total_silver": round(total_silver,3),
        "total_alloy": round(total_alloy,3)
    })

@app.route("/add-gold-stock", methods=["POST"])
def add_gold_stock():

    data = request.get_json()

    new_stock = GoldStock(
        stock_type="gold",   # ✅ IMPORTANT
        gold_type=data.get("gold_type"),
        purity=data.get("purity"),
        weight=data.get("weight"),
        pieces=data.get("pieces")
    )

    db.session.add(new_stock)
    db.session.commit()

    return jsonify({"status": "success"})

@app.route("/get-gold-stock")
def get_gold_stock():

    page = request.args.get("page", 1, type=int)
    per_page = 6

    pagination = GoldStock.query.filter_by(
        stock_type="gold"
    ).order_by(GoldStock.id.desc()).paginate(page=page, per_page=per_page)

    items = pagination.items

    stock = []

    for i in items:
        stock.append({
            "id": i.id,
            "gold_type": i.gold_type,
            "purity": i.purity,
            "weight": i.weight,
            "pieces": i.pieces,
            "date": i.created_at.strftime("%d-%m-%Y")
        })

    return jsonify({
        "data": stock,
        "page": page,
        "total_pages": pagination.pages
    })
    
@app.route("/gold-stock-summary")
def gold_stock_summary():

    # ===============================
    # TOTAL GOLD ONLY
    # ===============================
    total_weight = db.session.query(
        db.func.sum(GoldStock.weight)
    ).filter(
        GoldStock.stock_type == "gold"
    ).scalar() or 0

    total_pieces = db.session.query(
        db.func.sum(GoldStock.pieces)
    ).filter(
        GoldStock.stock_type == "gold"
    ).scalar() or 0


    # ===============================
    # FUNCTION (REUSABLE 🔥)
    # ===============================
    def get_purity_data(purity):

        weight = db.session.query(
            db.func.sum(GoldStock.weight)
        ).filter(
            GoldStock.stock_type == "gold",
            GoldStock.purity == purity
        ).scalar() or 0

        pieces = db.session.query(
            db.func.sum(GoldStock.pieces)
        ).filter(
            GoldStock.stock_type == "gold",
            GoldStock.purity == purity
        ).scalar() or 0

        return round(weight, 3), pieces


    # ===============================
    # GET ALL PURITIES
    # ===============================
    weight24, pieces24 = get_purity_data("24K")
    weight22, pieces22 = get_purity_data("22K")
    weight18, pieces18 = get_purity_data("18K")
    weight14, pieces14 = get_purity_data("14K")
    weight9,  pieces9  = get_purity_data("9K")


    # ===============================
    # RESPONSE
    # ===============================
    return jsonify({

        "total_weight": round(total_weight, 3),
        "total_pieces": total_pieces,

        "weight24": weight24,
        "pieces24": pieces24,

        "weight22": weight22,
        "pieces22": pieces22,

        "weight18": weight18,
        "pieces18": pieces18,

        "weight14": weight14,
        "pieces14": pieces14,

        "weight9": weight9,
        "pieces9": pieces9
    })
    


@app.route("/search-client-suggestions", methods=["POST"])
def search_client_suggestions():

    data = request.get_json()
    term = data.get("term","")

    clients = (
        ClientKYC.query
        .filter(ClientKYC.client_code.like(f"%{term}%"))
        .limit(5)
        .all()
    )

    return jsonify([c.client_code for c in clients])

@app.route("/save-mou", methods=["POST"])
def save_mou():

    client_code = request.form.get("client_code")

    client = ClientKYC.query.filter_by(client_code=client_code).first()

    if not client:
        return jsonify({"status":"client_not_found"})

    mou = ClientMOU(

        client_id = client.id,

        gold_credit_days = request.form.get("gold_credit_days"),
        labour_credit_days = request.form.get("labour_credit_days"),
        diamond_credit_days = request.form.get("diamond_credit_days"),
        chain_credit_days = request.form.get("chain_credit_days"),
        certificate_hallmark_days = request.form.get("certificate_hallmark_days"),

        making_charges = request.form.get("making_charges"),
        chain_charges = request.form.get("chain_charges"),
        synthetic_stone_charges = request.form.get("synthetic_stone_charges"),

        minimum_labour_rule = request.form.get("minimum_labour_rule"),
        nose_pin_labour = request.form.get("nose_pin_labour"),

        gold_labour_percent = request.form.get("gold_labour_percent"),

        # ⭐ NEW VALUES
        diamond_clarity = request.form.get("diamond_clarity"),
        diamond_color = request.form.get("diamond_color"),

        diamond_rate_2 = request.form.get("diamond_rate_2"),
        diamond_rate_6_5 = request.form.get("diamond_rate_6_5"),
        diamond_rate_11 = request.form.get("diamond_rate_11")
    )

    file = request.files.get("mou_document")

    if file:
        mou.mou_document = file.read()

    db.session.add(mou)
    db.session.commit()

    return jsonify({"status":"success"})

@app.route("/search-client-kyc", methods=["POST"])
def search_client_kyc():

    data = request.get_json()
    client_code = data.get("client_code")

    client = ClientKYC.query.filter_by(client_code=client_code).first()

    if not client:
        return jsonify({"status": "not_found"})

    owners = ClientOwner.query.filter_by(client_id=client.id).all()

    owner_list = []

    for o in owners:
        owner_list.append({
            "id": o.id,   # ⭐ REQUIRED FOR OWNER DOCUMENTS
            "full_name": o.full_name,
            "pan_number": o.pan_number,
            "aadhar_number": o.aadhar_number,
            "phone_number": o.phone_number,
            "email_id": o.email_id,
            "address": o.address
        })

    return jsonify({
        "status": "success",

        "client_code": client.client_code,
        "company_name": client.company_name,
        "gst_number": client.gst_number,
        "company_pan": client.company_pan,
        "msme": client.msme_registration,
        "iec_code": client.iec_code,
        "address": client.address,

        "bank_name": client.bank_name,
        "account_number": client.account_number,
        "ifsc_code": client.ifsc_code,
        "branch": client.branch,

        "owners": owner_list
    })
    

    
@app.route("/fetch-client-mou", methods=["POST"])
def fetch_client_mou():

    data = request.get_json()
    client_code = data.get("client_code")

    client = ClientKYC.query.filter_by(client_code=client_code).first()

    if not client:
        return jsonify({"status": "client_not_found"})

    mou = ClientMOU.query.filter_by(client_id=client.id).first()

    if not mou:
        return jsonify({"status": "mou_not_found"})

    return jsonify({

        "status": "success",

        # Credit days
        "gold_credit_days": mou.gold_credit_days,
        "labour_credit_days": mou.labour_credit_days,
        "diamond_credit_days": mou.diamond_credit_days,
        "chain_credit_days": mou.chain_credit_days,
        "certificate_hallmark_days": mou.certificate_hallmark_days,

        # Charges
        "making_charges": float(mou.making_charges or 0),
        "chain_charges": float(mou.chain_charges or 0),
        "synthetic_stone_charges": float(mou.synthetic_stone_charges or 0),

        # Rules
        "minimum_labour_rule": mou.minimum_labour_rule,
        "nose_pin_labour": mou.nose_pin_labour,
        "gold_labour_percent": float(mou.gold_labour_percent or 0),

        # ⭐ Diamond fields
        "diamond_clarity": mou.diamond_clarity,
        "diamond_color": mou.diamond_color,

        "diamond_rate_2": float(mou.diamond_rate_2 or 0),
        "diamond_rate_6_5": float(mou.diamond_rate_6_5 or 0),
        "diamond_rate_11": float(mou.diamond_rate_11 or 0),

        # MOU document
        "mou_view_url": f"/view-mou/{client.id}"
    })
    
@app.route("/view-mou/<int:client_id>")
def view_mou(client_id):

    mou = ClientMOU.query.filter_by(client_id=client_id).first()

    if not mou or not mou.mou_document:
        return "MOU Document Not Found"

    return Response(
        mou.mou_document,
        mimetype="application/pdf"
    )
    
@app.route("/view-doc/<doc_type>/<client_code>")
def view_doc(doc_type, client_code):

    client = ClientKYC.query.filter_by(client_code=client_code).first()

    if not client:
        return "Client Not Found", 404

    file_data = None

    if doc_type == "gst":
        file_data = client.gst_doc

    elif doc_type == "pan":
        file_data = client.pan_doc

    elif doc_type == "msme":
        file_data = client.msme_doc

    elif doc_type == "iec":
        file_data = client.iec_doc

    elif doc_type == "aadhaar":
        file_data = client.aadhaar_doc

    elif doc_type == "bank":
        file_data = client.bank_doc

    if not file_data:
        return "No File", 404

    # Detect file type
    if file_data.startswith(b"%PDF"):
        mime = "application/pdf"
    else:
        mime = "image/jpeg"

    return Response(file_data, mimetype=mime)

@app.route("/view-owner-doc/<doc_type>/<int:owner_id>")
def view_owner_doc(doc_type, owner_id):

    owner = ClientOwner.query.get(owner_id)

    if not owner:
        return "Owner not found", 404

    file_data = None

    if doc_type == "pan":
        file_data = owner.pan_doc

    elif doc_type == "aadhaar":
        file_data = owner.aadhar_doc

    if not file_data:
        return "No file", 404


    # =========================
    # Detect File Type
    # =========================

    if file_data.startswith(b"%PDF"):
        mime = "application/pdf"

    else:
        img_type = imghdr.what(None, file_data)

        if img_type == "png":
            mime = "image/png"
        elif img_type == "jpeg":
            mime = "image/jpeg"
        else:
            mime = "image/jpeg"


    return Response(
        file_data,
        mimetype=mime,
        headers={
            "Content-Disposition": "inline"
        }
    )
    

def generate_batch():
    return f"BATCH{random.randint(1000,9999)}"

@app.route("/api/production/update-section", methods=["POST"])
def update_production_section():

    data = request.get_json()

    raw_order = data.get("order_id")  # "SJOD0003, SJOD0002"
    department = data.get("department")
    section_data = data.get("data")

    if not raw_order or not department:
        return jsonify({"ok": False, "error": "Missing data"})

    order_nos = [o.strip() for o in raw_order.split(",")]

    orders = ProductionOrder.query.filter(
        ProductionOrder.order_no.in_(order_nos)
    ).all()

    if not orders:
        return jsonify({"ok": False, "error": "Orders not found"})

    # 🔥 create ONE batch
    batch_no = generate_batch()

    for order in orders:

        record = ProductionStatus.query.filter_by(order_id=order.id).first()

        if not record:
            record = ProductionStatus(
                order_id=order.id,
                batch_no=batch_no,
                data={}
            )
            db.session.add(record)

        else:
            record.batch_no = batch_no  # update batch if needed

        # update JSON
        current_data = record.data or {}
        current_data[department] = section_data
        record.data = current_data

        flag_modified(record, "data")

    db.session.commit()

    return jsonify({
        "ok": True,
        "batch_no": batch_no
    })


@app.route("/api/production/status-by-batch/<batch_no>", methods=["GET"])
def get_status_by_batch(batch_no):

    records = ProductionStatus.query.filter_by(batch_no=batch_no).all()

    if not records:
        return jsonify({"ok": False, "data": {}})

    # take first (same data for all)
    return jsonify({
        "ok": True,
        "data": records[0].data
    })






@app.route("/api/style-image/<style_no>", methods=["GET"])
def api_style_image(style_no):
    row = (
        ImageCrop.query
        .filter(ImageCrop.style_no == style_no)
        .order_by(ImageCrop.id.desc())
        .first()
    )

    if not row or not row.image:
        return ("", 404)

    img_bytes = row.image

    # If you always store JPEG, keep image/jpeg
    # If you store PNG, change to image/png
    return Response(img_bytes, mimetype="image/jpeg")


@app.route("/api/sieve-lookup-batch", methods=["POST"])
def sieve_lookup_batch():
    payload = request.get_json(silent=True) or {}
    mm_list = payload.get("mm_list", [])

    if not isinstance(mm_list, list) or not mm_list:
        return jsonify({})

    out = {}

    for raw_mm in mm_list:
        key = str(raw_mm)
    
        try:
            mm = round(float(raw_mm), 2)
        except:
            out[key] = None
            continue
    
        # ✅ Step 1: Near exact match
        row = (
            DiamondSieveMaster.query
            .filter(
                and_(
                    DiamondSieveMaster.mm_size >= mm - 0.001,
                    DiamondSieveMaster.mm_size <= mm + 0.001
                )
            )
            .order_by(DiamondSieveMaster.id.asc())
            .first()
        )
    
        # ✅ Step 2: fallback (floor)
        if not row:
            row = (
                DiamondSieveMaster.query
                .filter(DiamondSieveMaster.mm_size < mm)
                .order_by(
                    DiamondSieveMaster.mm_size.desc(),
                    DiamondSieveMaster.id.asc()
                )
                .first()
            )

        if not row:
            out[key] = None
            continue

        out[key] = {
            "mm_input": mm,
            "mm_size": float(row.mm_size),
            "sieve_range": row.sieve_range,
            "sieve_size": row.sieve_size,
            "ct_weight_per_piece": float(row.ct_weight_per_piece),
            "picked_id": int(row.id),
        }

    return jsonify(out)



def extract_cell_images(file_bytes):
    images = {}

    with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as z:

        # Read relationships
        rel_root = ET.fromstring(z.read("xl/_rels/cellimages.xml.rels"))

        rid_map = {}

        for rel in rel_root.findall("rel:Relationship", NS):
            rid_map[rel.attrib["Id"]] = rel.attrib["Target"]

        # Read image mapping
        cell_root = ET.fromstring(z.read("xl/cellimages.xml"))

        for cell in cell_root.findall("etc:cellImage", NS):

            pic = cell.find("xdr:pic", NS)
            if pic is None:
                continue

            cNvPr = pic.find("xdr:nvPicPr/xdr:cNvPr", NS)
            if cNvPr is None:
                continue

            image_name = cNvPr.attrib.get("name")

            blip = pic.find("xdr:blipFill/a:blip", NS)
            if blip is None:
                continue

            rid = blip.attrib.get("{%s}embed" % NS["r"])

            target = rid_map.get(rid)

            if not target:
                continue

            data = z.read("xl/" + target)

            images[image_name] = data

    return images


def get_embedded_image(cell_formula, images_map):

    if not isinstance(cell_formula, str):
        return None

    m = DISPIMG_RE.search(cell_formula)

    if not m:
        return None

    image_id = m.group(1)

    return images_map.get(image_id)


@app.route("/api/style-excel-upload", methods=["POST"])
def style_excel_upload():
    auth = require_admin_json()
    if auth:
        return auth

    f = request.files.get("file")
    clear_first = request.form.get("clear_first") == "1"

    if not f:
        return jsonify({"error": "No file uploaded"}), 400

    filename = secure_filename(f.filename or "")
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Upload only .xlsx files"}), 400

    file_bytes = f.read()

    # Read embedded images
    try:
        images_map = extract_cell_images(file_bytes)
    except Exception as e:
        images_map = {}
        app.logger.warning(f"Image extraction failed: {e}")

    # Open workbook
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"Excel read failed: {str(e)}"}), 400

    # Read header
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

    required = [
        "style_no",
        "category",
        "gold_wt",
        "net_wt",
        "dia_wt",
        "dia_pc",
        "cstn_pc",
        "cstn_wt",
        "brand"
    ]

    missing = [col for col in required if col not in headers]

    if missing:
        return jsonify({
            "error": f"Missing columns: {', '.join(missing)}"
        }), 400

    col_idx = {name: idx for idx, name in enumerate(headers)}

    if clear_first:
        db.session.query(StyleNo).delete()
        db.session.commit()

    total_rows = 0
    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2):

        total_rows += 1

        values = [cell.value for cell in row]

        def get(col_name):
            idx = col_idx.get(col_name)
            if idx is None or idx >= len(values):
                return None
            return values[idx]

        style_no = str(get("style_no") or "").strip()

        if not style_no:
            skipped += 1
            continue

        category = str(get("category") or "").strip()

        try:
            gold_wt = float(get("gold_wt")) if get("gold_wt") not in (None, "") else None
        except:
            gold_wt = None

        try:
            net_wt = float(get("net_wt")) if get("net_wt") not in (None, "") else None
        except:
            net_wt = None

        try:
            dia_wt = float(get("dia_wt")) if get("dia_wt") not in (None, "") else None
        except:
            dia_wt = None

        try:
            dia_pc = int(get("dia_pc")) if get("dia_pc") not in (None, "") else None
        except:
            dia_pc = None
            
        try:
            cstn_pc = int(get("cstn_pc")) if get("cstn_pc") not in ("", None) else None
        except:
            cstn_pc = None
        
        try:
            cstn_wt = float(get("cstn_wt")) if get("cstn_wt") not in ("", None) else None
        except:
            cstn_wt = None

        brand = str(get("brand") or "").strip()

        image_data = get_embedded_image(get("image"), images_map)
        gem_chart_data = get_embedded_image(get("gem_chart"), images_map)

        style = StyleNo(
            style_no=style_no,
            category=category,
            gold_wt=gold_wt,
            net_wt=net_wt,
            dia_wt=dia_wt,
            dia_pc=dia_pc,
            cstn_pc=cstn_pc,
            cstn_wt=cstn_wt,
            brand=brand,
            image=image_data,
            gem_chart=gem_chart_data
        )

        db.session.add(style)
        inserted += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "error": f"Database insert failed: {str(e)}"
        }), 500

    return jsonify({
        "message": "Style Excel imported successfully.",
        "total_rows": total_rows,
        "inserted": inserted,
        "skipped": skipped
    }), 200


def require_admin_json():
    if "role" not in session or session["role"] != "admin":
        return jsonify({"error": "Unauthorized. Please login as admin."}), 401
    return None


@app.route("/admin/upload-style-and-images")
def upload_style_and_images():
    if "role" not in session or session["role"] != "admin":
        return redirect("/signin")
    return render_template("upload_style_and_images.html")

# ---------- Client Code Generator ----------

def generate_client_code():

    last_client = ClientKYC.query.order_by(ClientKYC.id.desc()).first()

    if not last_client or not last_client.client_code:
        return "SJCC0001"

    last_number = int(last_client.client_code.replace("SJCC", ""))

    new_number = last_number + 1

    return f"SJCC{new_number:04d}"


# ---------- Upload Docs Route ----------

@app.route("/upload-docs", methods=["POST"])
def upload_docs():

    def file_to_bytes(file):
        if file and file.filename != "":
            return file.read()
        return None

    # =====================
    # GENERATE CLIENT CODE
    # =====================
    client_code = generate_client_code()

    # =====================
    # SAVE CLIENT KYC
    # =====================
    kyc = ClientKYC(

        client_code=client_code,

        # BUSINESS
        company_name=request.form.get("company_name"),
        gst_number=request.form.get("gst_number"),
        company_pan=request.form.get("company_pan"),
        msme_registration=request.form.get("msme_registration"),
        iec_code=request.form.get("iec_code"),
        address=request.form.get("address"),

        # BANK
        bank_name=request.form.get("bank_name"),
        account_number=request.form.get("account_number"),
        ifsc_code=request.form.get("ifsc_code"),
        branch=request.form.get("branch"),

        # DOCUMENTS
        pan_doc=file_to_bytes(request.files.get("pan_doc")),
        gst_doc=file_to_bytes(request.files.get("gst_doc")),
        msme_doc=file_to_bytes(request.files.get("msme_doc")),
        iec_doc=file_to_bytes(request.files.get("iec_doc")),
        visiting_card_doc=file_to_bytes(request.files.get("visiting_card_doc")),
        bank_doc=file_to_bytes(request.files.get("bank_doc")),
    )

    db.session.add(kyc)
    db.session.commit()

    # =====================
    # SAVE OWNERS
    # =====================

    full_names = request.form.getlist("full_name[]")
    pan_numbers = request.form.getlist("pan_number[]")
    aadhar_numbers = request.form.getlist("aadhar_number[]")
    phone_numbers = request.form.getlist("phone_number[]")
    emails = request.form.getlist("email_id[]")
    addresses = request.form.getlist("address[]")

    pan_docs = request.files.getlist("pan_doc[]")
    aadhar_docs = request.files.getlist("aadhar_doc[]")

    for i in range(len(full_names)):

        owner = ClientOwner(

            client_id=kyc.id,

            full_name=full_names[i] if i < len(full_names) else None,
            pan_number=pan_numbers[i] if i < len(pan_numbers) else None,
            aadhar_number=aadhar_numbers[i] if i < len(aadhar_numbers) else None,
            phone_number=phone_numbers[i] if i < len(phone_numbers) else None,
            email_id=emails[i] if i < len(emails) else None,
            address=addresses[i] if i < len(addresses) else None,

            pan_doc=file_to_bytes(pan_docs[i]) if i < len(pan_docs) else None,
            aadhar_doc=file_to_bytes(aadhar_docs[i]) if i < len(aadhar_docs) else None,
        )

        db.session.add(owner)

    db.session.commit()

    return jsonify({
        "status": "success",
        "message": "Client KYC submitted successfully",
        "client_code": kyc.client_code,
        "client_id": kyc.id
    })
    


@app.route("/portal", methods=["GET"])
def portal():
    return render_template("portal.html")



@app.route("/")
def home():
    return render_template("home.html")
    


@app.route("/sales")
def sales_page():
    if "role" not in session or session["role"] != "sales":
        return redirect("/signin")
    return render_template("sales.html")

@app.route("/acc_orders")
def acc_orders_page():
    if "role" not in session or session["role"] not in ["account", "admin"]:
        return redirect("/signin")
    

    orders = ProductionOrder.query.options(
        joinedload(ProductionOrder.client),
        joinedload(ProductionOrder.items)
    ).order_by(ProductionOrder.id.desc()).all()

    return render_template("acc_orders.html", orders=orders)

@app.route("/create_mou")
def create_mou():
    return render_template("create_mou.html")

@app.route("/mou_list")
def mou_list():
    return render_template("mou_list.html")


@app.route("/admin")
def admin():
    if "role" not in session or session["role"] != "admin":
        return redirect("/signin")
    return render_template("admin.html")

@app.route("/gold_items")
def gold_items():
    if "role" not in session or session["role"] != "admin":
        return redirect("/signin")
    return render_template("gold_items.html")

@app.route("/creat_emp")
def creat_emp():
    if "role" not in session or session["role"] != "admin":
        return redirect("/signin")
    return render_template("creat_emp.html")

@app.route("/kyc_list")
def kyc_list():
    if "role" not in session or session["role"] != "admin":
        return redirect("/signin")
    return render_template("kyc_list.html")

@app.route("/orders_status")
def orders_status():
    if "role" not in session or session["role"] != "admin":
        return redirect("/signin")
    return render_template("orders_status.html",)


@app.route("/styles")
def styles():
    if "role" not in session or session["role"] != "admin":
        return redirect("/signin")

    page = request.args.get("page", 1, type=int)
    per_page = 10  # ✅ change to 10/20/50

    # ✅ get only style_nos for this page (distinct)
    style_nos = (
        db.session.query(StyleNo.style_no)
        .filter(StyleNo.style_no.isnot(None))
        .group_by(StyleNo.style_no)
        .order_by(db.func.max(StyleNo.id).desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    style_nos = [s[0] for s in style_nos if s[0]]

    if not style_nos:
        return render_template("styles.html", styles=[], page=page, per_page=per_page, has_next=False)

    # ✅ fetch size rows only for these styles
    size_rows = (
        StyleNo.query
        .filter(StyleNo.style_no.in_(style_nos))
        .order_by(StyleNo.style_no.asc(), StyleNo.id.asc())
        .all()
    )

    # ✅ fetch images only for these styles (still heavy, but limited to 20)
    images = (
        ImageCrop.query
        .filter(ImageCrop.style_no.in_(style_nos))
        .all()
    )

    img_map = {}
    for im in images:
        if im.style_no not in img_map and im.image:
            img_map[im.style_no] = base64.b64encode(im.image).decode("utf-8")

    grouped = defaultdict(lambda: {
        "style_no": "",
        "gold_wt": None,
        "img_b64": None,
        "size_rows": []
    })

    for r in size_rows:
        g = grouped[r.style_no]
        g["style_no"] = r.style_no
        if g["gold_wt"] is None and r.gold_wt is not None:
            g["gold_wt"] = r.gold_wt

        if g["img_b64"] is None and r.style_no in img_map:
            g["img_b64"] = img_map[r.style_no]

        g["size_rows"].append({"size": r.size, "pcs": r.pcs})

    # keep same order as style_nos
    styles = [grouped[s] for s in style_nos if s in grouped]

    # ✅ check if next page exists
    next_exists = (
        db.session.query(StyleNo.style_no)
        .group_by(StyleNo.style_no)
        .order_by(db.func.max(StyleNo.id).desc())
        .offset(page * per_page)
        .limit(1)
        .count()
    ) > 0

    return render_template("styles.html", styles=styles, page=page, per_page=per_page, has_next=next_exists)

@app.route("/signup")
def signup():
    return render_template("signup.html")

# -----------------------------
# PAGES ROUTES (ERP MODULES)
# -----------------------------


@app.route("/create_order")
def create_order():
    clients = ClientKYC.query.order_by(ClientKYC.id.desc()).all()
    return render_template("create_order.html", clients=clients)

@app.route("/Client_kyc")
def Client_kyc():
    return render_template("Client_kyc.html")


@app.route("/production-board")
def production_board():
    if "role" not in session or session["role"] != "production":
        return redirect("/signin")

    orders = ProductionOrder.query.order_by(ProductionOrder.id.desc()).all()
    return render_template("production_board.html", orders=orders)




@app.route("/upload-doc", methods=["POST"])
def upload_doc():
    file = request.files.get("file")
    doc_type = request.form.get("doc_type")

    if not file:
        return jsonify({})

    result = process_image_ocr(file, doc_type)
    return jsonify(result)

@app.route("/upload-pdf-doc", methods=["POST"])
def upload_pdf_doc():
    file = request.files.get("file")
    doc_type = request.form.get("doc_type")

    if not file:
        return jsonify({})

    pdf_bytes = file.read()
    pdf = fitz.open(stream=pdf_bytes, filetype="pdf")

    if pdf.page_count == 0:
        return jsonify({})

    page = pdf.load_page(0)
    pix = page.get_pixmap(dpi=300)

    img_bytes = pix.tobytes("png")
    img_io = BytesIO(img_bytes)

    image_file = FileStorage(
        stream=img_io,
        filename="page1.png",
        content_type="image/png"
    )

    # ✅ ONLY place jsonify here
    result = process_image_ocr(image_file, doc_type)
    return jsonify(result)


# ---------- Total ocr Functions ----------

def process_image_ocr(file, doc_type):

    if not file:
        return {}

    # ---------- IMAGE READ ----------
    image = Image.open(file).convert("RGB")
    img = np.array(image)

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bitwise_not(gray)
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)

    # ---------- OCR ----------
    text = pytesseract.image_to_string(
        thresh,
        config="--oem 3 --psm 6"
    )

    print("===== OCR TEXT =====")
    print(text)
    print("====================")

    clean_text = text.upper()

    # ---------- FIX COMMON OCR ERRORS ----------
    clean_text = (
        clean_text
        .replace("GMAILL", "GMAIL")
        .replace("GMAILLC", "GMAIL")
        .replace("GMAIL.C0M", "GMAIL.COM")
    )

    response = {}

    # ================= AADHAR =================
    if doc_type == "aadhar":
        match = re.search(r"\d[\d\s]{11,14}", clean_text)
        if match:
            aadhar = re.sub(r"\s+", "", match.group())
            if len(aadhar) == 12:
                response["aadhar_number"] = aadhar

    # ================= PAN =================
    elif doc_type == "pan":
    
        lines = [l.strip() for l in clean_text.split("\n") if l.strip()]
        pan_text = re.sub(r"[^A-Z0-9]", "", clean_text)
    
        # ---------- PAN NUMBER ----------
        pan_text = text.replace(" ", "")
        
        for i in range(len(pan_text) - 9):
            chunk = pan_text[i:i+10]
            fixed = ""
    
            for idx, ch in enumerate(chunk):
                if idx < 5:
                    ch = ch.replace("0", "O").replace("1", "I")
                    if not ch.isalpha():
                        break
                elif idx < 9:
                    ch = ch.replace("O", "0").replace("I", "1")
                    if not ch.isdigit():
                        break
                else:
                    ch = ch.replace("0", "O").replace("1", "I")
                    if not ch.isalpha():
                        break
    
                fixed += ch
            else:
                response["pan_number"] = fixed
                break
    
        # ---------- NAME ----------
        for i, line in enumerate(lines):
            if "NAME" in line.upper():
    
                if i + 1 < len(lines):
                    name = lines[i + 1]
    
                    # Remove non alphabet
                    name = re.sub(r"[^A-Z ]", "", name.upper())
    
                    # Remove extra spaces
                    words = name.split()
    
                    # Keep only first 2 or 3 words (PAN names usually short)
                    if len(words) >= 2:
                        response["full_name"] = " ".join(words[:2])
    
                break
    


    # ================= MSME / UDYAM =================
    elif doc_type == "msme":
    
        # ---------- UDYAM NUMBER ----------
        udyam_match = re.search(
            r"UDYAM[-\s]*[A-Z]{2}[-\s]*\d{2}[-\s]*\d{7}",
            clean_text
        )
        if udyam_match:
            response["udyam_number"] = udyam_match.group().replace(" ", "")
    
        # ---------- ENTERPRISE NAME ----------
        lines = clean_text.splitlines()
    
        for i, line in enumerate(lines):
            if "NAME OF ENTERPRISE" in line:
    
                same_line = line.replace("NAME OF ENTERPRISE", "").strip()
    
                if len(same_line) > 3:
                    name = re.sub(r"[^A-Z ]", "", same_line)
                    name = re.sub(r"\s+", " ", name).strip()
                    response["enterprise_name"] = name.title()
    
                elif i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    name = re.sub(r"[^A-Z ]", "", next_line)
                    name = re.sub(r"\s+", " ", name).strip()
                    response["enterprise_name"] = name.title()
    
                break
    
        # ---------- MOBILE NUMBER ----------
        mobile_match = re.search(r"\b[6-9]\d{9}\b", clean_text)
        if mobile_match:
            response["mobile"] = mobile_match.group()
    
        # ---------- EMAIL ----------
        email_match = re.search(
            r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}",
            clean_text
        )
        if email_match:
            response["email"] = email_match.group()
            
# ================= IEC =================
    elif doc_type == "iec":
    
        # ---------- IEC CODE ----------
        iec_match = re.search(
            r"\b[A-Z]{5}[0-9]{4}[A-Z]\b",
            clean_text
        )
        if iec_match:
            response["iec_code"] = iec_match.group()
    
        # ---------- FIRM NAME ----------
        lines = clean_text.splitlines()
        for i, line in enumerate(lines):
    
            if "FIRM NAME" in line or "NAME OF FIRM" in line:
                same_line = (
                    line.replace("FIRM NAME", "")
                        .replace("NAME OF FIRM", "")
                        .strip()
                )
    
                # CASE 1: Name on same line
                if len(same_line) > 3:
                    name = re.sub(r"[^A-Z ]", "", same_line)
                    name = re.sub(r"\s+", " ", name).strip()
                    response["firm_name"] = name.title()
    
                # CASE 2: Name on next line
                elif i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    name = re.sub(r"[^A-Z ]", "", next_line)
                    name = re.sub(r"\s+", " ", name).strip()
                    response["firm_name"] = name.title()
    
                break
    # ================= GST =================
    elif doc_type == "gst":
    
        # ---------- CLEAN OCR TEXT ----------
        clean_text_upper = clean_text.upper()
        clean_text_no_space = re.sub(r"[\s]", "", clean_text_upper)
    
        # ---------- GST NUMBER ----------
        gst_pattern = r"\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]Z[A-Z0-9]"
        gst_match = re.search(gst_pattern, clean_text_no_space)
        
        if gst_match:
            response["gst_number"] = gst_match.group()
        
        
        # ---------- ADDRESS ----------
        lines = clean_text.splitlines()
        
        address_lines = []
        capture = False
        
        for line in lines:
            line_clean = line.strip()
            line_upper = line_clean.upper()
        
            # Detect address label
            if "ADDRESS OF PRINCIPAL PLACE" in line_upper:
        
                capture = True
        
                # Remove label text and keep remaining address
                part = re.split(r'ADDRESS OF PRINCIPAL PLACE OF', line_clean, flags=re.I)
        
                if len(part) > 1:
                    remaining = part[1].strip()
                    if remaining:
                        address_lines.append(remaining)
        
                continue
        
            # Stop when next section starts
            if capture and (
                "DATE OF LIABILITY" in line_upper or
                "DATE OF VALIDITY" in line_upper or
                "TYPE OF REGISTRATION" in line_upper
            ):
                break
        
            if capture and line_clean:
                address_lines.append(line_clean)
        
        
        if address_lines:
            address = " ".join(address_lines)
            address = re.sub(r"\s+", " ", address)
            response["address"] = address
            
    # ================= COMPANY PAN =================
    elif doc_type == "comp_pan":
    
        lines = [l.strip() for l in clean_text.split("\n") if l.strip()]
        pan_text = re.sub(r"[^A-Z0-9]", "", clean_text)
    
        # ---------- PAN NUMBER ----------
        for i in range(len(pan_text) - 9):
            chunk = pan_text[i:i+10]
            fixed = ""
    
            for idx, ch in enumerate(chunk):
    
                if idx < 5:
                    ch = ch.replace("0", "O").replace("1", "I")
                    if not ch.isalpha():
                        break
    
                elif idx < 9:
                    ch = ch.replace("O", "0").replace("I", "1")
                    if not ch.isdigit():
                        break
    
                else:
                    ch = ch.replace("0", "O").replace("1", "I")
                    if not ch.isalpha():
                        break
    
                fixed += ch
    
            else:
                response["pan_number"] = fixed
                break
    
    
        # ---------- COMPANY NAME ----------
        for i, line in enumerate(lines):
        
            if response.get("pan_number") and response["pan_number"] in line:
        
                if i + 1 < len(lines):
        
                    name = lines[i + 1].strip()
        
                    # remove special characters
                    name = re.sub(r"[^A-Za-z0-9&.,()\- ]", "", name)
        
                    # remove words like NONE
                    name = re.sub(r"\bNONE\b", "", name, flags=re.IGNORECASE)
        
                    # remove leading garbage (A, 2A, etc.)
                    name = re.sub(r"^[A-Z0-9]{1,3}\s+", "", name)
        
                    name = name.strip()
        
                    if len(name) > 5:
                        response["company_name"] = name
                break
                
        
    return response



# ---------- Get Client Names ----------
@app.route("/api/clients")
def api_clients():
    rows = ClientKYC.query.with_entities(
        ClientKYC.id,
        ClientKYC.full_name
    ).all()

    return jsonify([
        {"id": r.id, "full_name": r.full_name}
        for r in rows
    ])









def parse_dt(s):
    if not s:
        return None
    # "YYYY-MM-DDTHH:MM"
    return datetime.fromisoformat(s)

def generate_order_no():

    last = ProductionOrder.query.order_by(ProductionOrder.id.desc()).first()

    if not last or not last.order_no:
        return "SJOD0001"

    last_number = int(last.order_no.replace("SJOD", ""))
    new_number = last_number + 1

    return f"SJOD{new_number:04d}"

@app.route("/api/create-order", methods=["POST"])
def api_create_order():

    data = request.get_json(silent=True) or {}

    client_id = int(data.get("client_id") or 0)
    order_dt = parse_dt(data.get("order_datetime"))
    delivery_dt = parse_dt(data.get("delivery_datetime"))

    items = data.get("items") or []

    if not client_id:
        return jsonify({"ok": False, "error": "client_id required"}), 400

    if not order_dt:
        return jsonify({"ok": False, "error": "order_datetime required"}), 400

    if not items:
        return jsonify({"ok": False, "error": "At least one item required"}), 400


    # GENERATE ORDER NUMBER
    order_no = generate_order_no()


    order = ProductionOrder(
        order_no=order_no,
        client_id=client_id,
        order_datetime=order_dt,
        delivery_datetime=delivery_dt,
        total_amount=float(data.get("total_amount") or 0),
        remark=(data.get("remark") or "").strip()
    )

    db.session.add(order)
    db.session.flush()


    for item in items:

        purity_factor = item.get("gold_purity_factor")
        purity_factor = float(purity_factor) if purity_factor else None

        row = ProductionOrderItem(
            order_id=order.id,

            style_no=(item.get("style_no") or "").strip(),

            diamond_clarity=item.get("diamond_clarity"),
            gold_color=item.get("gold_color"),
            diamond_color=item.get("diamond_color"),

            gold_purity=item.get("gold_purity"),
            gold_purity_factor=purity_factor,

            pieces=int(item.get("pieces") or 1),
            style_remark=item.get("style_remark")
        )

        db.session.add(row)

    db.session.commit()

    return jsonify({
        "ok": True,
        "order_id": order.id,
        "order_no": order.order_no
    }), 201
    
    





@app.route("/signin", methods=["GET", "POST"])
def signin():
    selected_role = request.args.get("role")

    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        user = User.query.filter_by(email=email, password=password).first()

        if user:
            session["user_id"] = user.id
            session["role"] = user.role

            if user.role == "admin":
                return redirect("/admin")
            elif user.role == "sales":
                return redirect("/sales")
            elif user.role == "production":
                return redirect("/production-board")
            elif user.role == "account":
                return redirect("/acc_orders")

        return "Invalid Login"

    return render_template("signin.html", selected_role=selected_role)



@app.route("/logout")
def logout():
    session.clear()
    return redirect("/portal")




# -----------------------------
# RUN SERVER
# -----------------------------

if __name__ == "__main__":
    app.run(debug=True)
